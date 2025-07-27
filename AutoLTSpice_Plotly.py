import subprocess
import os
import sys
import shutil
import itertools
from datetime import datetime
import webbrowser

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

import plotly.graph_objects as go
from plotly.subplots import make_subplots

import numpy as np

# --- Utility Functions ---

def edit_asc_file(asc_path: str, edit_func) -> None:
    with open(asc_path, 'r') as f:
        lines = f.readlines()
    lines = edit_func(lines)
    with open(asc_path, 'w') as f:
        f.writelines(lines)


def run_ltspice(asc_path: str, ltspice_exe: str) -> None:
    try:
        subprocess.run([ltspice_exe, '-b', asc_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=True)
    except KeyboardInterrupt:
        print("Simulation aborted by user")
        sys.exit(1)
    except FileNotFoundError:
        print(f"LTspice not found at: {ltspice_exe}")
        sys.exit(1)
    except subprocess.CalledProcessError as e:
        print(f"LTspice returned code {e.returncode}")
        sys.exit(1)


def extract_results(txt_path: str) -> dict:
    result = {"Efficiency": None, "Vratio": None, "Vripple": None, "THD": None}
    if not os.path.isfile(txt_path):
        print(f"Log file not found: {txt_path}")
        sys.exit(1)
    with open(txt_path, 'r') as f:
        for line in f:
            lower = line.lower()
            if "efficiency:" in lower:
                parts = line.strip().split("=")
                if len(parts) == 2:
                    try:
                        result["Efficiency"] = float(parts[1].strip())
                    except:
                        pass
            if "vratio:" in lower:
                parts = line.strip().split("=")
                if len(parts) == 2:
                    try:
                        result["Vratio"] = float(parts[1].strip())
                    except:
                        pass
            if "vripple:" in lower:
                parts = line.strip().split("=")
                if len(parts) == 2:
                    try:
                        result["Vripple"] = float(parts[1].strip())
                    except:
                        pass
            if "total harmonic distortion:" in lower:
                thd_val = line.split(":", 1)[1].strip()
                try:
                    result["THD"] = float(thd_val.replace("%", ""))
                except:
                    pass
    return result


def generic_edit(lines: list, param_dict: dict) -> list:
    for idx, line in enumerate(lines):
        for param, value in param_dict.items():
            if f".param {param} =" in line or f"!.param {param} =" in line:
                prefix, _, suffix = line.partition(f"{param} =")
                after = suffix.lstrip()
                rest = after.partition(" ")[2] if " " in after else ""
                lines[idx] = f"{prefix}{param} = {value} {rest}".rstrip() + "\n"
    return lines


def add_table_header(ws, text, start_col, end_col, row):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col)
    cell.value = text
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def run_suite(asc, log, names, params, label, ltspice) -> pd.DataFrame:
    results = []
    combos = list(itertools.product(*params))
    total = len(combos)
    for idx, combo in enumerate(combos, 1):
        pdict = dict(zip(names, combo))
        if 'D1' in pdict and 'D2' in pdict:
            d1, d2 = pdict['D1'], pdict['D2']
            if not (d1 == 0 or d2 == 0 or d1 == d2):
                continue
        print(f"Iteration {idx}/{total} ({label}): {pdict}")
        edit_asc_file(asc, lambda L: generic_edit(L, pdict))
        run_ltspice(asc, ltspice)
        vals = extract_results(log)
        results.append({**pdict,
                        'Efficiency': vals['Efficiency'],
                        'Vratio': vals['Vratio'],
                        'Vripple': vals['Vripple'],
                        'THD(%)': vals['THD']})
    return pd.DataFrame(results)


if __name__ == "__main__":
    # Parameter ranges
    vin_vals   = [60]#list(range(10, 121, 15))
    freq_vals  = [50000, 100000, 150000, 200000, 250000, 300000]
    DCR_vals   = [0.01]
    induc_vals = [10e-6]
    induc1_vals= [5e-6]
    induc2_vals= [5e-6]
    cap1_vals  = [10e-6]
    cap2_vals  = [47e-6]
    D1_vals    = list(np.round(np.arange(0, 0.96, 0.1), 2))
    D2_vals    = list(np.round(np.arange(0, 0.96, 0.1), 2))

    # Config paths
    buck_asc, buck_log = (
        r"C:\Users\adam2\OneDrive\Documents\SolarEdge\SolarEdge LTSpice\BuckBoost_Test2.asc",
        r"C:\Users\adam2\OneDrive\Documents\SolarEdge\SolarEdge LTSpice\BuckBoost_Test2.log"
    )
    boost_asc, boost_log = (
        r"C:\Users\adam2\OneDrive\Documents\SolarEdge\SolarEdge LTSpice\BoostBuckTest2.asc",
        r"C:\Users\adam2\OneDrive\Documents\SolarEdge\SolarEdge LTSpice\BoostBuckTest2.log"
    )

    for path in (buck_asc, boost_asc):
        if not os.path.isfile(path):
            print(f"Error: .asc file not found: {path}")
            sys.exit(1)
    default_exe = r"C:\Program Files\LTS\LTspice.exe"
    ltspice = default_exe if os.path.isfile(default_exe) else (shutil.which("XVIIx64.exe") or shutil.which("LTspice.exe"))
    if not ltspice:
        print("Error: LTspice executable not found.")
        sys.exit(1)

    # Run simulations
    buck_df = run_suite(
        buck_asc, buck_log,
        ["Vin","freq","DCR_val","induc_val","cap1_val","cap2_val","D1","D2"],
        [vin_vals, freq_vals, DCR_vals, induc_vals, cap1_vals, cap2_vals, D1_vals, D2_vals],
        label="buck", ltspice=ltspice
    )
    boost_df = run_suite(
        boost_asc, boost_log,
        ["Vin","freq","DCR_val","induc1_val","induc2_val","cap1_val","cap2_val","D1","D2"],
        [vin_vals, freq_vals, DCR_vals, induc1_vals, induc2_vals, cap1_vals, cap2_vals, D1_vals, D2_vals],
        label="boost", ltspice=ltspice
    )

    # Prepare for Excel
    def prepare_for_excel(df, mode='buck'):
        df_ex = df.copy()
        df_ex['freq']    = df_ex['freq'] / 1e3   # kHz
        df_ex['DCR_val'] = df_ex['DCR_val']*1e3  # mΩ
        df_ex['Vripple'] = df_ex['Vripple']*100  # %
        if mode=='buck': df_ex['induc_val']=df_ex['induc_val']*1e6
        else:
            df_ex['induc1_val']=df_ex['induc1_val']*1e6
            df_ex['induc2_val']=df_ex['induc2_val']*1e6
        df_ex['cap1_val']=df_ex['cap1_val']*1e6
        df_ex['cap2_val']=df_ex['cap2_val']*1e6
        rename_map={
            'Vin':'Vin (V)', 'freq':'freq (kHz)', 'DCR_val':'DCR (mΩ)',
            'induc_val':'L (µH)','induc1_val':'L1 (µH)','induc2_val':'L2 (µH)',
            'cap1_val':'C1 (µF)','cap2_val':'C2 (µF)',
            'D1':'D1','D2':'D2','Efficiency':'Efficiency',
            'Vratio':'Vratio','Vripple':'Vripple (%)','THD(%)':'THD (%)'
        }
        return df_ex.rename(columns=rename_map)

    buck_df_excel=prepare_for_excel(buck_df,'buck')
    boost_df_excel=prepare_for_excel(boost_df,'boost')

    ts=datetime.now().strftime("%Y%m%d_%H%M%S")
    base_dir=os.path.join(os.path.expanduser("~"),'OneDrive','Documents','SolarEdge','SolarEdge_DualSimExcel')
    os.makedirs(base_dir,exist_ok=True)
    excel_path=os.path.join(base_dir,f"SolarEdge_Results_{ts}.xlsx")
    with pd.ExcelWriter(excel_path,engine='openpyxl') as writer:
        buck_df_excel.to_excel(writer,sheet_name='Results',index=False,startrow=2)
        boost_df_excel.to_excel(writer,sheet_name='Results',index=False,startrow=2,startcol=len(buck_df_excel.columns)+3)
    wb=openpyxl.load_workbook(excel_path)
    ws=wb['Results']
    add_table_header(ws,'Buck-Boost',1,len(buck_df_excel.columns),2)
    add_table_header(ws,'Boost-Buck',len(buck_df_excel.columns)+4,len(buck_df_excel.columns)*2+4,2)
    start_row=3; end_row=start_row+len(buck_df_excel)
    buck_tbl=Table(displayName='TableBuck',ref=f"A{start_row}:{get_column_letter(len(buck_df_excel.columns))}{end_row}")
    boost_tbl=Table(displayName='TableBoost',ref=f"{get_column_letter(len(buck_df_excel.columns)+4)}{start_row}:{get_column_letter(len(buck_df_excel.columns)*2+4)}{end_row}")
    for tbl in (buck_tbl,boost_tbl):
        tbl.tableStyleInfo=TableStyleInfo(name='TableStyleMedium9',showFirstColumn=False,showRowStripes=True)
        ws.add_table(tbl)
    wb.save(excel_path)
    try: os.startfile(excel_path)
    except: pass
    
### Final Unified Plotly Section: Correct Delta Placement Fix
from itertools import cycle
from plotly.colors import DEFAULT_PLOTLY_COLORS
import glob, os, webbrowser
from plotly.subplots import make_subplots
import plotly.graph_objects as go

buck_df_plot  = buck_df[~((buck_df['D1']==0) & (buck_df['D2']==0))]
boost_df_plot = boost_df[~((boost_df['D1']==0) & (boost_df['D2']==0))]

# Initialize consistent color cycle
color_cycle = cycle(DEFAULT_PLOTLY_COLORS)
color_map = {}

def get_color(name):
    if name not in color_map:
        color_map[name] = next(color_cycle)
    return color_map[name]

metrics = [
    ('Efficiency', 'Efficiency'),
    ('Vripple', 'Ripple Voltage (%)'),
    ('Vratio', 'Voltage Ratio'),
    ('THD(%)', 'THD (%)')
]

# --- Replace your old “# Build pivot tables” block with this ---
# Prepare filtered pivots for D1-axis (filter D2 extremes)
pivots_buck_filtered   = {}
pivots_boost_filtered  = {}
# Prepare filtered pivots for D2-axis (filter D1 extremes)
pivots_buck_filtered_D2  = {}
pivots_boost_filtered_D2 = {}
# Prepare filtered pivots for diagonal (D1==D2)
pivots_buck_diag       = {}
pivots_boost_diag      = {}

for key, _ in metrics:
    # D1-axis pivots (D2 extremes)
    df_b1  = buck_df_plot[buck_df_plot['D2'].isin([0, 1])]
    df_bo1 = boost_df_plot[boost_df_plot['D2'].isin([0, 1])]
    pivots_buck_filtered[key]  = df_b1 .pivot_table(index='D1', columns='freq', values=key, aggfunc='mean')
    pivots_boost_filtered[key] = df_bo1.pivot_table(index='D1', columns='freq', values=key, aggfunc='mean')

    # D2-axis pivots (D1 extremes)
    df_b2  = buck_df_plot[buck_df_plot['D1'].isin([0, 1])]
    df_bo2 = boost_df_plot[boost_df_plot['D1'].isin([0, 1])]
    pivots_buck_filtered_D2[key]  = df_b2 .pivot_table(index='D2', columns='freq', values=key, aggfunc='mean')
    pivots_boost_filtered_D2[key] = df_bo2.pivot_table(index='D2', columns='freq', values=key, aggfunc='mean')

    # Diagonal pivots (D1 == D2)
    df_bd   = buck_df_plot[buck_df_plot['D1'] == buck_df_plot['D2']]
    df_bo_bd= boost_df_plot[boost_df_plot['D1'] == boost_df_plot['D2']]
    pivots_buck_diag[key]  = df_bd   .pivot_table(index='D1', columns='freq', values=key, aggfunc='mean')
    pivots_boost_diag[key] = df_bo_bd.pivot_table(index='D1', columns='freq', values=key, aggfunc='mean')

os.makedirs('charts', exist_ok=True)

# 1) Combined metrics (2 rows x 3 cols)
for key, label in metrics:
    # build 2×3 grid: data on top, deltas below
    fig = make_subplots(
        rows=2, cols=3,
        subplot_titles=[
            f'{label} (Boost Mode)',
            f'{label} (Buck Mode)',
            f'{label} (Buck-Boost Mode)',
            f'Buck Minus Boost (Boost Mode)',
            f'Buck Minus Boost (Buck Mode)',
            f'Buck Minus Boost (Buck-Boost Mode)'
        ]
    )
    
    # three main sections: D1, D2, D1=D2
    sections = [
        (pivots_buck_filtered[key], pivots_boost_filtered[key]),
        (pivots_buck_filtered_D2[key], pivots_boost_filtered_D2[key]),
        (pivots_buck_diag[key], pivots_boost_diag[key])
    ]
    axes = ['D1', 'D2', 'D1 = D2']

    for idx, (sec, ax_label) in enumerate(zip(sections, axes), start=1):
        bb, bo = sec
        # if we’re plotting ripple, convert fraction→percent
        if key == 'Vripple':
            bb = bb * 100
            bo = bo * 100

        # top row: combined Buck-Boost & Boost-Buck
        for freq in bb.columns:
            c_bb = get_color(f'Buck-Boost {int(freq/1e3)} kHz')
            c_bo = get_color(f'Boost-Buck  {int(freq/1e3)} kHz')
            fig.add_trace(
                go.Scatter(x=bb.index, y=bb[freq], mode='markers+lines',
                           name=f'Buck-Boost {int(freq/1e3)} kHz',
                           marker=dict(color=c_bb), line=dict(color=c_bb)),
                row=1, col=idx
            )
            fig.add_trace(
                go.Scatter(x=bo.index, y=bo[freq], mode='markers+lines',
                           name=f'Boost-Buck {int(freq/1e3)} kHz',
                           marker=dict(color=c_bo), line=dict(color=c_bo)),
                row=1, col=idx
            )
        fig.update_xaxes(title_text=ax_label, row=1, col=idx)
        fig.update_yaxes(title_text=label, row=1, col=idx)

        # bottom row: delta = bb - bo
        delta = bb.subtract(bo, fill_value=0)
        for freq in delta.columns:
            c_d = get_color(f'{int(freq/1e3)} kHz')
            fig.add_trace(
                go.Scatter(x=delta.index, y=delta[freq], mode='markers+lines',
                           name=f'{int(freq/1e3)} kHz',
                           marker=dict(color=c_d), line=dict(color=c_d)),
                row=2, col=idx
            )
        fig.update_xaxes(title_text=f'{ax_label}', row=2, col=idx)
        fig.update_yaxes(title_text=f'Δ {label}', row=2, col=idx)

    # clean up legend & save
    seen = set()
    for t in fig.data:
        if t.name in seen:
            t.showlegend = False
        else:
            seen.add(t.name)
    fig.update_layout(title_text=label, showlegend=True, height=900, width=1400)
    path = os.path.join('charts', f"{key.replace('%','pct')}_combined.html")
    fig.write_html(path)
    webbrowser.open(path)


# 2) Vratio-specific plots (2 rows x 3 cols)
fig2 = make_subplots(
        rows=2, cols=3,
        subplot_titles=[
            f'Vratio vs Efficiency (Boost Mode)',
            f'Vratio vs Efficiency (Buck Mode)',
            f'Vratio vs Efficiency (Buck-Boost Mode)',
            f'Buck Minus Boost (Boost Mode)',
            f'Buck Minus Boost (Buck Mode)',
            f'Buck Minus Boost (Buck-Boost Mode)'
        ]
    )

sections2 = [
    (pivots_buck_filtered['Vratio'], pivots_buck_filtered['Efficiency'],
     pivots_boost_filtered['Vratio'], pivots_boost_filtered['Efficiency']),
    (pivots_buck_filtered_D2['Vratio'], pivots_buck_filtered_D2['Efficiency'],
     pivots_boost_filtered_D2['Vratio'], pivots_boost_filtered_D2['Efficiency']),
    (pivots_buck_diag['Vratio'], pivots_buck_diag['Efficiency'],
     pivots_boost_diag['Vratio'], pivots_boost_diag['Efficiency'])
]

for i, (vp, ep, bv, be) in enumerate(sections2, start=1):
    # Raw Vratio vs Efficiency
    for freq in vp.columns:
        name_bb = f'Buck-Boost {int(freq/1e3)} kHz'
        name_bo = f'Boost-Buck {int(freq/1e3)} kHz'
        color_bb = get_color(name_bb)
        color_bo = get_color(name_bo)
        fig2.add_trace(
            go.Scatter(
                x=vp[freq], y=ep[freq], mode='markers+lines',
                name=name_bb, marker=dict(color=color_bb), line=dict(color=color_bb)
            ), row=1, col=i
        )
        fig2.add_trace(
            go.Scatter(
                x=bv[freq], y=be[freq], mode='markers+lines',
                name=name_bo, marker=dict(color=color_bo), line=dict(color=color_bo)
            ), row=1, col=i
        )
    fig2.update_xaxes(title_text='Vratio', row=1, col=i)
    fig2.update_yaxes(title_text='Efficiency', row=1, col=i)

     # Nearest-point Δ Efficiency
    for freq in vp.columns:
        x_buck = vp[freq].values
        y_buck = ep[freq].values
        x_boost = bv[freq].values
        y_boost = be[freq].values
        x_diff = []
        y_diff = []
        for xb, yb in zip(x_buck, y_buck):
            idx_nearest = (np.abs(x_boost - xb)).argmin()
            x_diff.append(xb)
            y_diff.append(yb - y_boost[idx_nearest])
        name_d = f'{int(freq/1e3)} kHz'
        color_d = get_color(name_d)
        fig2.add_trace(
            go.Scatter(
                x=x_diff, y=y_diff, mode='markers+lines', name=name_d,
                marker=dict(color=color_d), line=dict(color=color_d)
            ), row=2, col=i
        )
    fig2.update_xaxes(title_text='Vratio (Nearest Match)', row=2, col=i)
    fig2.update_yaxes(title_text='Δ Efficiency', row=2, col=i)

# Deduplicate legend entries for fig2
seen2 = set()
for trace in fig2.data:
    if trace.name in seen2:
        trace.showlegend = False
    else:
        seen2.add(trace.name)

fig2.update_layout(
    title_text='Vratio vs Efficiency Combined',
    showlegend=True,
    height=900, width=1400
)
out2 = os.path.join('charts','Vratio_vs_Efficiency_combined.html')
fig2.write_html(out2)
webbrowser.open_new_tab(f"file://{os.path.abspath(out2)}")

# Section 3: Frequency-specific Vratio vs Efficiency (6-panel per freq)
for freq in pivots_buck_filtered['Vratio'].columns:
    sections = [
        ('Boost Mode', pivots_buck_filtered['Vratio'], pivots_buck_filtered['Efficiency'],
         pivots_boost_filtered['Vratio'], pivots_boost_filtered['Efficiency']),
        ('Buck Mode', pivots_buck_filtered_D2['Vratio'], pivots_buck_filtered_D2['Efficiency'],
         pivots_boost_filtered_D2['Vratio'], pivots_boost_filtered_D2['Efficiency']),
        ('Buck-Boost Mode', pivots_buck_diag['Vratio'], pivots_buck_diag['Efficiency'],
         pivots_boost_diag['Vratio'], pivots_boost_diag['Efficiency'])
    ]

    fig6 = make_subplots(
        rows=2, cols=3,
        subplot_titles=[
            f'Vratio vs Efficiency ({mode})' for mode, *_ in sections
        ] + [
            f'Buck Minus Boost ({mode})' for mode, *_ in sections
        ]
    )

    delta_name = f'{int(freq/1e3)} kHz'
    c_d = get_color(delta_name)

    for col, (mode, vr_all, eff_all, bvr_all, bef_all) in enumerate(sections, start=1):
        # Top row: raw Buck-Boost vs Boost-Buck
        name_bb = f'Buck-Boost {int(freq/1e3)} kHz'
        name_bo = f'Boost-Buck {int(freq/1e3)} kHz'
        c_bb = get_color(name_bb)
        c_bo = get_color(name_bo)

        fig6.add_trace(
            go.Scatter(x=vr_all[freq], y=eff_all[freq], mode='markers+lines',
                       name=name_bb, marker=dict(color=c_bb), line=dict(color=c_bb)),
            row=1, col=col
        )
        fig6.add_trace(
            go.Scatter(x=bvr_all[freq], y=bef_all[freq], mode='markers+lines',
                       name=name_bo, marker=dict(color=c_bo), line=dict(color=c_bo)),
            row=1, col=col
        )
        fig6.update_xaxes(title_text='Vratio', row=1, col=col)
        fig6.update_yaxes(title_text='Efficiency', row=1, col=col)

        # Bottom row: nearest match delta
        x_diff = []
        y_diff = []
        vr_vals = vr_all[freq].values
        eff_vals = eff_all[freq].values
        bvr_vals = bvr_all[freq].values
        bef_vals = bef_all[freq].values
        for xb, yb in zip(vr_vals, eff_vals):
            idx_nearest = (np.abs(bvr_vals - xb)).argmin()
            x_diff.append(xb)
            y_diff.append(yb - bef_vals[idx_nearest])

        fig6.add_trace(
            go.Scatter(x=x_diff, y=y_diff, mode='markers+lines',
                       name=delta_name, marker=dict(color=c_d), line=dict(color=c_d)),
            row=2, col=col
        )
        fig6.update_xaxes(title_text='Vratio (Nearest Match)', row=2, col=col)
        fig6.update_yaxes(title_text='Δ Efficiency', row=2, col=col)

    # Clean up legend & finalize
    seen6 = set()
    for tr in fig6.data:
        if tr.name in seen6:
            tr.showlegend = False
        else:
            seen6.add(tr.name)

    fig6.update_layout(
        title_text=f'Vratio vs Efficiency @{int(freq/1e3)} kHz',
        showlegend=True, height=900, width=1400
    )
    out6 = os.path.join('charts', f"Vratio_vs_Efficiency_panels_{int(freq/1e3)}kHz.html")
    fig6.write_html(out6)
    webbrowser.open_new_tab(f"file://{os.path.abspath(out6)}")

# 4) Frequency-specific averaged plots (2 rows x 3 cols per freq)
for key, label in metrics:
    factor = 100 if key == 'Vripple' else 1
    for freq in buck_df_plot['freq'].unique():
        # Prepare averaged series by axis
        # Boost Mode: D2=0, group by D1
        df_bb = buck_df_plot[(buck_df_plot['freq']==freq) & (buck_df_plot['D2']==0)]
        df_bo = boost_df_plot[(boost_df_plot['freq']==freq) & (boost_df_plot['D2']==0)]
        ser_bb_D1 = df_bb.groupby('D1')[key].mean() * factor
        ser_bo_D1 = df_bo.groupby('D1')[key].mean() * factor
        # Buck Mode: D1=0, group by D2
        df_bb2 = buck_df_plot[(buck_df_plot['freq']==freq) & (buck_df_plot['D1']==0)]
        df_bo2 = boost_df_plot[(boost_df_plot['freq']==freq) & (boost_df_plot['D1']==0)]
        ser_bb_D2 = df_bb2.groupby('D2')[key].mean() * factor
        ser_bo_D2 = df_bo2.groupby('D2')[key].mean() * factor
        # Buck-Boost Mode: D1==D2, group by D1=D2
        df_bd = buck_df_plot[(buck_df_plot['freq']==freq) & (buck_df_plot['D1']==buck_df_plot['D2'])]
        df_bo_bd = boost_df_plot[(boost_df_plot['freq']==freq) & (boost_df_plot['D1']==boost_df_plot['D2'])]
        ser_bb_diag = df_bd.groupby('D1')[key].mean() * factor
        ser_bo_diag = df_bo_bd.groupby('D1')[key].mean() * factor

        # Build subplots
        fig = make_subplots(
            rows=2, cols=3,
            subplot_titles=[
                f'{label} (Boost Mode)', f'{label} (Buck Mode)', f'{label} (Buck-Boost Mode)',
                f'Buck Minus Boost (Boost Mode)', f'Buck Minus Boost (Buck Mode)', f'Buck Minus Boost (Buck-Boost Mode)'
            ]
        )
        # Top row: raw averaged curves
        # Col 1: D1 axis
        fig.add_trace(go.Scatter(x=ser_bb_D1.index, y=ser_bb_D1.values, mode='lines+markers', name='Buck-Boost', line=dict(color=get_color(f'BB-{freq}')), marker=dict(color=get_color(f'BB-{freq}'))), row=1, col=1)
        fig.add_trace(go.Scatter(x=ser_bo_D1.index, y=ser_bo_D1.values, mode='lines+markers', name='Boost-Buck', line=dict(color=get_color(f'BO-{freq}')), marker=dict(color=get_color(f'BO-{freq}'))), row=1, col=1)
        # Col 2: D2 axis
        fig.add_trace(go.Scatter(x=ser_bb_D2.index, y=ser_bb_D2.values, mode='lines+markers', name='Buck-Boost', line=dict(color=get_color(f'BB-{freq}')), marker=dict(color=get_color(f'BB-{freq}'))), row=1, col=2)
        fig.add_trace(go.Scatter(x=ser_bo_D2.index, y=ser_bo_D2.values, mode='lines+markers', name='Boost-Buck', line=dict(color=get_color(f'BO-{freq}')), marker=dict(color=get_color(f'BO-{freq}'))), row=1, col=2)
        # Col 3: diagonal
        fig.add_trace(go.Scatter(x=ser_bb_diag.index, y=ser_bb_diag.values, mode='lines+markers', name='Buck-Boost', line=dict(color=get_color(f'BB-{freq}')), marker=dict(color=get_color(f'BB-{freq}'))), row=1, col=3)
        fig.add_trace(go.Scatter(x=ser_bo_diag.index, y=ser_bo_diag.values, mode='lines+markers', name='Boost-Buck', line=dict(color=get_color(f'BO-{freq}')), marker=dict(color=get_color(f'BO-{freq}'))), row=1, col=3)

        # Axes labels
        fig.update_xaxes(title_text='D1', row=1, col=1)
        fig.update_xaxes(title_text='D2', row=1, col=2)
        fig.update_xaxes(title_text='D1 = D2', row=1, col=3)
        fig.update_yaxes(title_text=label, row=1, col=1)
        fig.update_yaxes(title_text=label, row=1, col=2)
        fig.update_yaxes(title_text=label, row=1, col=3)

        # Bottom row: delta curves
        delta1 = ser_bb_D1 - ser_bo_D1.reindex(ser_bb_D1.index, method='nearest')
        delta2 = ser_bb_D2 - ser_bo_D2.reindex(ser_bb_D2.index, method='nearest')
        deltad = ser_bb_diag - ser_bo_diag.reindex(ser_bb_diag.index, method='nearest')
        fig.add_trace(go.Scatter(x=delta1.index, y=delta1.values, mode='lines+markers', name=f'Δ {int(freq/1e3)} kHz', line=dict(color=get_color(f'D1-{freq}')), marker=dict(color=get_color(f'D1-{freq}'))), row=2, col=1)
        fig.add_trace(go.Scatter(x=delta2.index, y=delta2.values, mode='lines+markers', name=f'Δ {int(freq/1e3)} kHz', line=dict(color=get_color(f'D1-{freq}')), marker=dict(color=get_color(f'D1-{freq}'))), row=2, col=2)
        fig.add_trace(go.Scatter(x=deltad.index, y=deltad.values, mode='lines+markers', name=f'Δ {int(freq/1e3)} kHz', line=dict(color=get_color(f'D1-{freq}')), marker=dict(color=get_color(f'D1-{freq}'))), row=2, col=3)
        fig.update_xaxes(title_text='D1', row=2, col=1)
        fig.update_xaxes(title_text='D2', row=2, col=2)
        fig.update_xaxes(title_text='D1 = D2', row=2, col=3)
        fig.update_yaxes(title_text=f'Δ {label}', row=2, col=1)
        fig.update_yaxes(title_text=f'Δ {label}', row=2, col=2)
        fig.update_yaxes(title_text=f'Δ {label}', row=2, col=3)

        # clean up legend & save
        seen5 = set()
        for t in fig.data:
            if t.name in seen5:
                t.showlegend = False
            else:
                seen5.add(t.name)
        # Finalize
        fig.update_layout(title_text=f'{label} @ {int(freq/1e3)} kHz (Averaged)', showlegend=True, height=900, width=1200)
        out = os.path.join('charts', f"{key.replace('%','pct')}_{int(freq/1e3)}kHz_averaged.html")
        fig.write_html(out)
        webbrowser.open_new_tab(f"file://{os.path.abspath(out)}")

print('All plots completed.')
